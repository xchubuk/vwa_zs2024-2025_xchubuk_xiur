#basic imports of libs

from flask import Flask, render_template, request, jsonify, redirect, url_for, make_response, session, abort, send_file
from flask_wtf import CSRFProtect
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import tempfile
import bcrypt
import hashlib
import sqlite3
import uuid
import os
from datetime import datetime, timedelta
from dotenv import load_dotenv

#some global variables

load_dotenv()
SECRET_KEY = os.getenv('SECRET_KEY')
URL = os.getenv("URL")
PORT = os.getenv("PORT")

app = Flask(__name__)
app.secret_key = SECRET_KEY
csrf = CSRFProtect(app)

#extra helper functions

def generate_session_token(email):
    random_uuid = uuid.uuid4().hex
    token = hashlib.sha256(f'{email}{random_uuid}'.encode()).hexdigest()
    return token

def fetch_user_role_from_db():
    session_token = request.cookies.get('session_token')
    if not session_token:
        return None

    conn = sqlite3.connect('bicycle_rental.db')
    cursor = conn.cursor()
    cursor.execute('''
        SELECT roles.name FROM users
        JOIN user_roles ON users.user_id = user_roles.user_id
        JOIN roles ON user_roles.role_id = roles.role_id
        WHERE users.session_token = ?
    ''', (session_token,))
    role = cursor.fetchone()
    conn.close()
    
    return role[0] if role else None

def fetch_bicycles():
    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT b.bicycle_id, b.inventory_number, b.type, t.name AS type_name, t.description, bs.status
            FROM bicycles b
            JOIN bicycle_types t ON b.type_id = t.type_id
            LEFT JOIN bicycle_status bs ON b.bicycle_id = bs.bicycle_id
            WHERE bs.status_id = (SELECT MAX(status_id) FROM bicycle_status WHERE bicycle_id = b.bicycle_id)
        ''')
        
        rows = cursor.fetchall()

        bicycles = []
        for row in rows:
            bicycles.append({
                "bicycle_id": row[0],
                "inventory_number": row[1],
                "type": row[2],
                "type_name": row[3],
                "description": row[4],
                "status": row[5]
            })
        
        conn.close()
        return bicycles
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return []
        

def fetch_user_role_from_db():
    session_token = session.get('session_token')
    if not session_token:
        return None

    conn = sqlite3.connect('bicycle_rental.db')
    cursor = conn.cursor()
    cursor.execute('''
        SELECT roles.name FROM users
        JOIN user_roles ON users.user_id = user_roles.user_id
        JOIN roles ON user_roles.role_id = roles.role_id
        WHERE users.session_token = ? AND users.session_expiration > ?
    ''', (session_token, datetime.utcnow()))
    role = cursor.fetchone()
    conn.close()
    
    return role[0] if role else None

#global rest api handlers

@app.route('/api/get_user_role')
def get_user_role():
    role = fetch_user_role_from_db()
    if role:
        return jsonify({"role": role})
    return jsonify({"error": "Role not found"}), 401

@app.route('/api/types', methods=['GET'])
def get_types():
    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT type FROM bicycle_types")
        
        types = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        return jsonify(types)
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify([]), 500

@app.route('/api/bicycles', methods=['GET'])
def get_bicycles():
    bicycles = fetch_bicycles()
    return jsonify(bicycles)

@app.route('/client_dashboard')
def client_dashboard():
    role = fetch_user_role_from_db()
    if role != 'client':
        return "Forbidden", 403
    return render_template('client_dashboard.html')

@app.route('/manager_dashboard')
def manager_dashboard():
    role = fetch_user_role_from_db()
    if role != 'manager':
        return "Forbidden", 403
    return render_template('manager_dashboard.html')

@app.route('/admin_dashboard')
def admin_dashboard():
    role = fetch_user_role_from_db()
    if role != 'admin':
        return "Forbidden", 403
    return render_template('admin_dashboard.html')

@app.route('/mechanic_dashboard')
def mechanic_dashboard():
    role = fetch_user_role_from_db()
    if role != 'mechanic':
        return "Forbidden", 403
    return render_template('mechanic_dashboard.html')

@app.route('/api/rent', methods=['POST'])
def rent_bicycle():
    print("Session data:", dict(session))  
    data = request.get_json()
    user_id = session.get('user_id')
    bicycle_id = data.get('bikeId')
    hours = data.get('hours')
    payment_type = data.get('payment')

    if not user_id or not bicycle_id or not hours or not payment_type:
        return jsonify({"error": "All fields are required"}), 400

    start_date = datetime.utcnow()
    end_date = start_date + timedelta(hours=int(hours))

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO rentals (user_id, bicycle_id, start_date, end_date)
            VALUES (?, ?, ?, ?)
        ''', (user_id, bicycle_id, start_date, end_date))
        
        rental_id = cursor.lastrowid

        cursor.execute('''
            INSERT INTO transactions (rental_id, amount, payment_method, payment_date, payment_status)
            VALUES (?, ?, ?, ?, ?)
        ''', (rental_id, hours * 10, payment_type, start_date, "Paid"))

        transaction_id = cursor.lastrowid

        cursor.execute('''
            UPDATE rentals 
            SET transaction_id = ?
            WHERE rental_id = ?
        ''', (transaction_id, rental_id))

        cursor.execute('''
            UPDATE bicycle_status 
            SET status = '0', 
                user_id = ?, 
                comment = ?,
                inspection_date = ?
            WHERE bicycle_id = ?
        ''', (user_id, "Bicycle rented", start_date, bicycle_id))

        conn.commit()
        conn.close()
        
        return jsonify({"message": "Rental successful"}), 200
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to rent bicycle"}), 500

@app.route('/')
def entry():
    if request.cookies.get('session_token'):
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/register', methods=['POST'])
@csrf.exempt
def handle_register():
    data = request.get_json()
    name = data.get('name')
    email = data.get('email')
    password = data.get('password')
    registration_time = datetime.fromtimestamp(data.get('registrationTime') / 1000).strftime('%Y-%m-%d %H:%M:%S')

    names = name.split()
    first_name = names[0] if len(names) > 0 else ""
    last_name = names[1] if len(names) > 1 else ""
    
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        cursor.execute('INSERT INTO users (first_name, last_name, email, password, registration_date) VALUES (?, ?, ?, ?, ?)', (first_name, last_name, email, hashed_password, registration_time))
        
        user_id = cursor.lastrowid

        session_token = generate_session_token(email)
        session_expiration = datetime.utcnow() + timedelta(hours=24)
        
        cursor.execute('UPDATE users SET session_token = ?, session_expiration = ? WHERE user_id = ?', (session_token, session_expiration, user_id))

        role_name = "client"
        cursor.execute('SELECT role_id FROM roles WHERE name = ?', (role_name,))
        role = cursor.fetchone()

        if role:
            role_id = role[0]
            start_date = registration_time
            cursor.execute('INSERT INTO user_roles (user_id, role_id, start_date) VALUES (?, ?, ?)', (user_id, role_id, start_date))

        conn.commit()
        conn.close()
        
        session['session_token'] = session_token
        session['user_id'] = user_id

        response = make_response(jsonify({"message": "Registration successful", "redirect_url": url_for('index'), "success": True}))
        return response
    except sqlite3.IntegrityError:
        return jsonify({"message": "Email is already registered"}), 400

@app.route('/login', methods=['POST'])
@csrf.exempt
def handle_login():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')

    conn = sqlite3.connect('bicycle_rental.db')
    cursor = conn.cursor()
    cursor.execute('SELECT user_id, password FROM users WHERE email = ?', (email,))
    row = cursor.fetchone()

    if row and bcrypt.checkpw(password.encode('utf-8'), row[1]):
        user_id = row[0]
        session_token = generate_session_token(email)
        session_expiration = datetime.utcnow() + timedelta(hours=24)
        
        cursor.execute('UPDATE users SET session_token = ?, session_expiration = ? WHERE user_id = ?', (session_token, session_expiration, user_id))
        conn.commit()
        conn.close()
        
        session['session_token'] = session_token
        session['user_id'] = user_id

        response = make_response(jsonify({"message": "Login successful", "redirect_url": url_for('index'), "success": True}))
        return response
    else:
        conn.close()
        return jsonify({"message": "Invalid credentials"}), 401
    
@app.route('/index')
def index():
    return render_template('index.html')

@app.route('/logout')
def logout():
    session.clear()  
    response = make_response(redirect(url_for('entry')))
    response.delete_cookie('session_token')  
    return response

#admin rest api calls

@app.route('/api/admin/bicycles', methods=['GET'])
def get_admin_bicycles():
    role = fetch_user_role_from_db()
    
    
    if role not in ['admin', 'manager']:
        return abort(403)  

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        
        cursor.execute('''
            SELECT b.bicycle_id, b.inventory_number, b.type, t.name AS type_name, 
                   t.description, t.purchase_date, t.type AS bicycle_type, bs.inspection_date, 
                   bs.user_id, bs.comment, bs.status
            FROM bicycles b
            JOIN bicycle_types t ON b.type_id = t.type_id
            LEFT JOIN bicycle_status bs ON b.bicycle_id = bs.bicycle_id
            WHERE bs.status_id = (SELECT MAX(status_id) FROM bicycle_status WHERE bicycle_id = b.bicycle_id)
            ORDER BY b.bicycle_id
        ''')

        rows = cursor.fetchall()

        bicycles = []
        for row in rows:
            bicycles.append({
                "bicycle_id": row[0],
                "inventory_number": row[1],
                "type": row[2],
                "type_name": row[3],
                "description": row[4],
                "purchase_date": row[5],
                "bicycle_type": row[6],
                "inspection_date": row[7],
                "user_id": row[8],
                "comment": row[9],
                "status": row[10]
            })

        conn.close()
        return jsonify(bicycles), 200
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to fetch bicycles"}), 500

@app.route('/api/admin/bicycles', methods=['POST'])
def add_bicycle():
    data = request.get_json()

    inventory_number = data.get('inventory_number')
    type_id = data.get('type_id')
    comment = data.get('comment')
    status = data.get('status')  
    inspection_date = data.get('inspection_date')  

    if not inventory_number or type_id is None or status is None:
        return jsonify({"error": "Missing required fields"}), 400

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        cursor.execute('SELECT type FROM bicycle_types WHERE type_id = ?', (type_id,))
        type_row = cursor.fetchone()
        
        if not type_row:
            return jsonify({"error": "Invalid type_id"}), 400

        type_value = type_row[0]

        cursor.execute('''
            INSERT INTO bicycles (inventory_number, type_id, type)
            VALUES (?, ?, ?)
        ''', (inventory_number, type_id, type_value))

        bicycle_id = cursor.lastrowid

        cursor.execute('''
            INSERT INTO bicycle_status (bicycle_id, inspection_date, status, user_id, comment)
            VALUES (?, ?, ?, NULL, ?)
        ''', (bicycle_id, inspection_date, str(status), comment))

        conn.commit()
        conn.close()
        return jsonify({"message": "Bicycle added successfully"}), 201
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to add bicycle"}), 500
    
@app.route('/api/bicycle_types', methods=['GET'])
def get_bicycle_types():
    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()
        
        cursor.execute("SELECT type_id, name FROM bicycle_types")
        types = [{"type_id": row[0], "name": row[1]} for row in cursor.fetchall()]
        
        conn.close()
        return jsonify(types)
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify([]), 500
    
@app.route('/api/admin/users/<int:user_id>/role', methods=['POST'])
def update_user_role(user_id):
    role = fetch_user_role_from_db()
    if role != 'admin':
        return abort(403)

    data = request.get_json()
    new_role = data.get('role')

    if new_role not in ["client", "manager", "admin", "mechanic"]:
        return jsonify({"error": "Invalid role"}), 400

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        cursor.execute('SELECT role_id FROM roles WHERE name = ?', (new_role,))
        role_id_row = cursor.fetchone()
        if not role_id_row:
            return jsonify({"error": "Role not found"}), 404
        
        new_role_id = role_id_row[0]

        cursor.execute('''
            UPDATE user_roles
            SET role_id = ?, start_date = ?
            WHERE user_id = ?
        ''', (new_role_id, datetime.utcnow(), user_id))

        conn.commit()
        conn.close()

        return jsonify({"message": "User role updated successfully"}), 200
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to update user role"}), 500
    
@app.route('/api/admin/bicycles/<int:bicycle_id>/status', methods=['POST'])
def update_bicycle_status(bicycle_id):
    role = fetch_user_role_from_db()
    if role not in ['admin', 'manager']:
        return abort(403)
    
    data = request.get_json()
    new_status = data.get('status')
    comment = data.get('comment', '')
    user_id = session.get('user_id')  

    if new_status is None:
        return jsonify({"error": "Status is required"}), 400

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE bicycle_status 
            SET status = ?, user_id = ?, comment = ?, inspection_date = CURRENT_TIMESTAMP
            WHERE bicycle_id = ?
        ''', (str(new_status), user_id, comment, bicycle_id))

        conn.commit()
        conn.close()
        
        return jsonify({"message": "Bicycle status updated successfully"}), 200
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to update bicycle status"}), 500
    
@app.route('/api/admin/bicycles/<int:bicycle_id>', methods=['DELETE'])
def remove_bicycle(bicycle_id):
    role = fetch_user_role_from_db()
    if role != 'admin':
        return abort(403)

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM bicycle_status WHERE bicycle_id = ?', (bicycle_id,))
        cursor.execute('DELETE FROM bicycles WHERE bicycle_id = ?', (bicycle_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({"message": "Bicycle removed successfully"}), 200
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to remove bicycle"}), 500
    
@app.route('/api/admin/users', methods=['GET'])
def get_admin_users():
    role = fetch_user_role_from_db()
    if role != 'admin':
        return abort(403)

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT u.user_id, u.first_name, u.last_name, u.email, u.registration_date, r.name as role_name
            FROM users u
            LEFT JOIN user_roles ur ON u.user_id = ur.user_id
            LEFT JOIN roles r ON ur.role_id = r.role_id
            ORDER BY u.user_id
        ''')

        users = []
        for row in cursor.fetchall():
            users.append({
                "user_id": row[0],
                "first_name": row[1],
                "last_name": row[2],
                "email": row[3],
                "registration_date": row[4],
                "role": row[5]
            })

        conn.close()
        return jsonify(users)

    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to fetch users"}), 500
    
@app.route('/api/admin/add_user', methods=['POST'])
def add_user():
    role = fetch_user_role_from_db()
    if role != 'admin':
        return abort(403)  

    data = request.get_json()
    first_name = data.get('first_name')
    last_name = data.get('last_name')
    email = data.get('email')
    password = data.get('password')
    role_name = data.get('role')

    if not (first_name and last_name and email and password and role_name):
        return jsonify({"error": "All fields are required"}), 400

    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
    registration_time = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        
        cursor.execute('''
            INSERT INTO users (first_name, last_name, email, password, registration_date)
            VALUES (?, ?, ?, ?, ?)
        ''', (first_name, last_name, email, hashed_password, registration_time))
        
        user_id = cursor.lastrowid

        
        cursor.execute('SELECT role_id FROM roles WHERE name = ?', (role_name,))
        role = cursor.fetchone()
        if role:
            role_id = role[0]
            cursor.execute('''
                INSERT INTO user_roles (user_id, role_id, start_date)
                VALUES (?, ?, ?)
            ''', (user_id, role_id, registration_time))

        conn.commit()
        conn.close()
        
        return jsonify({"message": "User added successfully"}), 201
    except sqlite3.IntegrityError:
        return jsonify({"error": "Email is already registered"}), 400
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to add user"}), 500
    
@app.route('/api/admin/generate_report', methods=['GET'])
def generate_report():
    role = fetch_user_role_from_db()
    if role != 'admin':
        return abort(403)

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        cursor.execute('''
            SELECT 
                b.inventory_number,
                t.name AS type_name,
                bs.status,
                bs.inspection_date,
                r.start_date,
                r.end_date,
                u.first_name || ' ' || u.last_name as renter_name,
                tr.amount as rental_cost,
                tr.payment_method,
                tr.payment_status,
                rep.repair_notes,
                rep.completion_date as repair_date,
                m.first_name || ' ' || m.last_name as mechanic_name,
                sr.problem_description,
                sr.creation_date as repair_request_date
            FROM bicycles b
            JOIN bicycle_types t ON b.type_id = t.type_id
            LEFT JOIN bicycle_status bs ON b.bicycle_id = bs.bicycle_id
            LEFT JOIN rentals r ON b.bicycle_id = r.bicycle_id
            LEFT JOIN users u ON r.user_id = u.user_id
            LEFT JOIN transactions tr ON r.rental_id = tr.rental_id
            LEFT JOIN repairments rep ON b.bicycle_id = rep.bicycle_id
            LEFT JOIN users m ON rep.mechanic_id = m.user_id
            LEFT JOIN service_requests sr ON b.bicycle_id = sr.bicycle_id
            WHERE bs.status_id = (SELECT MAX(status_id) FROM bicycle_status WHERE bicycle_id = b.bicycle_id)
            ORDER BY b.inventory_number, r.start_date DESC
        ''')
        report_data = cursor.fetchall()
        conn.close()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Bicycle Report"

        headers = [
            "Inventory Number",
            "Type",
            "Current Status",
            "Last Inspection",
            "Last Rental Start",
            "Last Rental End",
            "Last Renter",
            "Rental Cost",
            "Payment Method",
            "Payment Status",
            "Last Repair Notes",
            "Last Repair Date",
            "Last Mechanic",
            "Problem Description",
            "Repair Request Date"
        ]
        sheet.append(headers)

        status_mapping = {
            "1": "Available",
            "0": "In Use",
            "-1": "Under Maintenance"
        }

        for row in report_data:
            inventory_number = row[0]
            type_name = row[1]
            status = status_mapping.get(str(row[2]), "Unknown")
            inspection_date = row[3].split(" ")[0] if row[3] else "-"
            rental_start = row[4].split(" ")[0] if row[4] else "-"
            rental_end = row[5].split(" ")[0] if row[5] else "Ongoing" if row[4] else "-"
            renter = row[6] if row[6] else "-"
            rental_cost = f"${row[7]}" if row[7] else "-"
            payment_method = row[8] if row[8] else "-"
            payment_status = row[9] if row[9] else "-"
            repair_notes = row[10] if row[10] else "-"
            repair_date = row[11].split(" ")[0] if row[11] else "-"
            mechanic = row[12] if row[12] else "-"
            problem_desc = row[13] if row[13] else "-"
            request_date = row[14].split(" ")[0] if row[14] else "-"

            sheet.append([
                inventory_number,
                type_name,
                status,
                inspection_date,
                rental_start,
                rental_end,
                renter,
                rental_cost,
                payment_method,
                payment_status,
                repair_notes,
                repair_date,
                mechanic,
                problem_desc,
                request_date
            ])

        for column in sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="009879", end_color="009879", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        workbook.save(temp_file.name)
        temp_file.seek(0)

        return send_file(
            temp_file.name, 
            as_attachment=True, 
            download_name=f"bicycle_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"Error generating report: {e}")
        return jsonify({"error": "Failed to generate report"}), 500

@app.route('/api/admin/delete_user/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    role = fetch_user_role_from_db()
    if role != 'admin':
        return abort(403)  

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM user_roles WHERE user_id = ?', (user_id,))
        cursor.execute('DELETE FROM users WHERE user_id = ?', (user_id,))
        
        conn.commit()
        conn.close()

        return jsonify({"message": "User deleted successfully"}), 200
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to delete user"}), 500
    
#manager rest api calls
    
@app.route('/api/manager/bicycles/<int:bicycle_id>/return', methods=['POST'])
def return_bicycle(bicycle_id):
    role = fetch_user_role_from_db()
    if role != 'manager':
        return abort(403)

    data = request.get_json()
    user_id = session.get('user_id')
    payment_method = data.get('payment_method', 'cash')

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        cursor.execute('''
            SELECT rental_id FROM rentals
            WHERE bicycle_id = ? AND end_date IS NULL
            ORDER BY start_date DESC
            LIMIT 1
        ''', (bicycle_id,))
        rental = cursor.fetchone()

        if rental:
            rental_id = rental[0]
            end_date = datetime.utcnow()

            cursor.execute('''
                UPDATE rentals
                SET end_date = ?
                WHERE rental_id = ?
            ''', (end_date, rental_id))

            cursor.execute('''
                INSERT INTO transactions (rental_id, amount, payment_method, payment_date, payment_status)
                VALUES (?, ?, ?, ?, ?)
            ''', (rental_id, 0, payment_method, end_date, "Completed"))

        cursor.execute('''
            UPDATE bicycle_status 
            SET status = '1', 
                user_id = ?, 
                comment = ?,
                inspection_date = CURRENT_TIMESTAMP
            WHERE bicycle_id = ?
        ''', (user_id, "Bicycle returned by manager", bicycle_id))

        conn.commit()
        conn.close()

        return jsonify({"message": "Bicycle returned successfully and status updated"}), 200
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to return bicycle"}), 500
    
@app.route('/api/manager/bicycles/<int:bicycle_id>/repair', methods=['POST'])
def repair_bicycle(bicycle_id):
    role = fetch_user_role_from_db()
    if role != 'manager':
        return abort(403)

    data = request.get_json()
    problem_description = data.get('problem_description')
    manager_id = session.get('user_id')

    if not problem_description:
        return jsonify({"error": "Problem description is required"}), 400

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO service_requests (bicycle_id, manager_id, problem_description, status, repaired, creation_date)
            VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ''', (bicycle_id, manager_id, problem_description, "Pending", False))

        cursor.execute('''
            UPDATE bicycle_status 
            SET status = '-1', 
                user_id = ?, 
                comment = ?,
                inspection_date = CURRENT_TIMESTAMP
            WHERE bicycle_id = ?
        ''', (manager_id, "Bicycle marked for repair", bicycle_id))

        conn.commit()
        conn.close()

        return jsonify({"message": "Repair request created and bicycle status updated"}), 200
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to submit repair request"}), 500
    
#mechanic rest api calls

@app.route('/api/mechanic/bicycles', methods=['GET'])
def get_mechanic_bicycles():
    role = fetch_user_role_from_db()
    if role != 'mechanic':
        return abort(403)

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        cursor.execute('''
            SELECT b.bicycle_id, b.inventory_number, b.type, t.name AS type_name, 
                   t.description, t.purchase_date, t.type AS bicycle_type,
                   sr.problem_description, sr.status AS repair_status,
                   r.completion_date as last_repair_date,
                   r.repair_notes as last_repair_notes
            FROM bicycles b
            JOIN bicycle_types t ON b.type_id = t.type_id
            JOIN bicycle_status bs ON b.bicycle_id = bs.bicycle_id
            LEFT JOIN service_requests sr ON b.bicycle_id = sr.bicycle_id
            LEFT JOIN (
                SELECT bicycle_id, completion_date, repair_notes
                FROM repairments
                WHERE completion_date = (
                    SELECT MAX(completion_date)
                    FROM repairments r2
                    WHERE r2.bicycle_id = repairments.bicycle_id
                )
            ) r ON b.bicycle_id = r.bicycle_id
            WHERE bs.status = '-1'
            AND bs.status_id = (
                SELECT MAX(status_id) 
                FROM bicycle_status 
                WHERE bicycle_id = b.bicycle_id
            )
            AND (sr.status = 'Pending' OR sr.status IS NULL)
            ORDER BY sr.creation_date DESC
        ''')

        columns = [column[0] for column in cursor.description]
        bicycles = [dict(zip(columns, row)) for row in cursor.fetchall()]

        conn.close()
        return jsonify(bicycles), 200

    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to fetch bicycles"}), 500

@app.route('/api/mechanic/bicycles/<int:bicycle_id>/repair-complete', methods=['POST'])
def complete_repair(bicycle_id):
    role = fetch_user_role_from_db()
    if role != 'mechanic':
        return abort(403)

    data = request.get_json()
    repair_notes = data.get('repair_notes')
    mechanic_id = session.get('user_id')

    if not repair_notes:
        return jsonify({"error": "Repair notes are required"}), 400

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO repairments (completion_date, mechanic_id, bicycle_id, repair_notes)
            VALUES (CURRENT_TIMESTAMP, ?, ?, ?)
        ''', (mechanic_id, bicycle_id, repair_notes))

        cursor.execute('''
            UPDATE service_requests 
            SET status = 'Completed', 
                closure_date = CURRENT_TIMESTAMP
            WHERE bicycle_id = ? 
            AND status = 'Pending'
        ''', (bicycle_id,))

        cursor.execute('''
            UPDATE bicycle_status 
            SET status = '1', 
                user_id = ?, 
                comment = ?,
                inspection_date = CURRENT_TIMESTAMP
            WHERE bicycle_id = ?
        ''', (mechanic_id, f"Repair completed: {repair_notes}", bicycle_id))

        conn.commit()
        conn.close()

        return jsonify({"message": "Repair completed successfully"}), 200

    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to complete repair"}), 500
    
#client rest api calls

@app.route('/api/client/rental_history', methods=['GET'])
def get_rental_history():
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({"error": "User not authenticated"}), 401

    try:
        conn = sqlite3.connect('bicycle_rental.db')
        cursor = conn.cursor()

        
        cursor.execute('''
            SELECT b.inventory_number, b.type, t.name AS bike_type, r.start_date, r.end_date
            FROM rentals r
            JOIN bicycles b ON r.bicycle_id = b.bicycle_id
            JOIN bicycle_types t ON b.type_id = t.type_id
            WHERE r.user_id = ?
            ORDER BY r.start_date DESC
        ''', (user_id,))
        
        rentals = []
        for row in cursor.fetchall():
            start_date = datetime.strptime(row[3], '%Y-%m-%d %H:%M:%S.%f')
            end_date = datetime.strptime(row[4], '%Y-%m-%d %H:%M:%S.%f') if row[4] else None
            duration = f"{(end_date - start_date).total_seconds() / 3600:.0f}h" if end_date else "Ongoing"

            rentals.append({
                "inventory_number": row[0],
                "bike_type": row[1],
                "start_date": start_date.strftime("%Y-%m-%d %H:%M"),
                "duration": duration
            })

        conn.close()
        return jsonify(rentals), 200
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Failed to fetch rental history"}), 500

if __name__ == '__main__':
    app.run(host=URL, port=PORT, debug=True)